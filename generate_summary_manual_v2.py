#!/usr/bin/env python3
"""
Manual Summary Generator v2 - With Per-Device Packet Loss Calculation + Excel Export

Usage:
    python3 generate_summary_manual_v2.py <path_to_mqtt_metrics_log.csv>

Example:
    python3 generate_summary_manual_v2.py /home/mqtt-sdn/results/06-dscp-qos-13switches/run_2025-11-24_13-36-51/mqtt_metrics_log.csv
"""

import sys
import csv
import os
from statistics import mean, stdev
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")
    print("Install with: pip3 install openpyxl")

def calculate_summary(csv_file):
    """Calculate metrics summary from CSV file with per-device packet loss"""

    # Per-type aggregated metrics
    delays = {'anomaly': [], 'normal': []}
    jitter_values = {'anomaly': [], 'normal': []}
    prev_delay = {'anomaly': None, 'normal': None}

    # Per-device tracking for packet loss and delays
    device_data = defaultdict(lambda: {
        'type': None,
        'received_seq': set(),
        'max_seq': -1,
        'delays': []
    })

    message_count = 0
    first_timestamp = None
    last_timestamp = None

    print(f"Reading CSV: {csv_file}")
    print("Processing messages...")

    with open(csv_file, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            device = row['device']
            msg_type = row['type']
            delay = float(row['delay_ms'])
            seq = int(row['seq'])
            timestamp = float(row['timestamp_sent'])

            # Aggregate delay
            delays[msg_type].append(delay)
            message_count += 1

            # Calculate jitter
            if prev_delay[msg_type] is not None:
                jitter = abs(delay - prev_delay[msg_type])
                jitter_values[msg_type].append(jitter)
            prev_delay[msg_type] = delay

            # Per-device sequence and delay tracking
            device_data[device]['type'] = msg_type
            device_data[device]['received_seq'].add(seq)
            device_data[device]['delays'].append(delay)
            if seq > device_data[device]['max_seq']:
                device_data[device]['max_seq'] = seq

            if first_timestamp is None:
                first_timestamp = timestamp
            last_timestamp = timestamp

            if message_count % 10000 == 0:
                print(f"  Processed {message_count} messages...")

    print(f"Total messages processed: {message_count}")
    print(f"Total unique devices: {len(device_data)}")

    # Calculate per-device packet loss, then aggregate by type
    loss_by_type = {'anomaly': [], 'normal': []}
    total_expected_by_type = {'anomaly': 0, 'normal': 0}
    total_received_by_type = {'anomaly': 0, 'normal': 0}
    total_lost_by_type = {'anomaly': 0, 'normal': 0}

    print("\nPer-device packet loss:")
    for device, data in sorted(device_data.items()):
        msg_type = data['type']
        max_seq = data['max_seq']
        expected = max_seq + 1  # seq starts from 0
        received = len(data['received_seq'])
        lost = expected - received
        loss_rate = (lost / expected * 100) if expected > 0 else 0

        loss_by_type[msg_type].append(loss_rate)
        total_expected_by_type[msg_type] += expected
        total_received_by_type[msg_type] += received
        total_lost_by_type[msg_type] += lost

        if loss_rate > 0:  # Only print devices with packet loss
            print(f"  {device:30s} ({msg_type:7s}): {received:5d}/{expected:5d} ({loss_rate:5.2f}% loss)")

    # Calculate summary
    duration = last_timestamp - first_timestamp if first_timestamp and last_timestamp else 0

    summary = {}
    for msg_type in ['anomaly', 'normal']:
        if delays[msg_type]:
            # Aggregate packet loss rate
            avg_loss_rate = mean(loss_by_type[msg_type]) if loss_by_type[msg_type] else 0

            summary[msg_type] = {
                'count': len(delays[msg_type]),
                'avg_delay': mean(delays[msg_type]),
                'min_delay': min(delays[msg_type]),
                'max_delay': max(delays[msg_type]),
                'std_delay': stdev(delays[msg_type]) if len(delays[msg_type]) > 1 else 0,
                'avg_jitter': mean(jitter_values[msg_type]) if jitter_values[msg_type] else 0,
                'num_devices': len([d for d, info in device_data.items() if info['type'] == msg_type]),
                'loss_info': {
                    'expected': total_expected_by_type[msg_type],
                    'received': total_received_by_type[msg_type],
                    'lost': total_lost_by_type[msg_type],
                    'loss_rate': (total_lost_by_type[msg_type] / total_expected_by_type[msg_type] * 100)
                                if total_expected_by_type[msg_type] > 0 else 0,
                    'avg_loss_per_device': avg_loss_rate
                }
            }

    summary['total'] = {
        'duration': duration,
        'total_messages': message_count,
        'throughput': message_count / duration if duration > 0 else 0
    }

    # Per-device statistics for detailed analysis
    summary['device_stats'] = []
    for device, data in device_data.items():
        if data['delays']:
            # Parse floor and room from device name
            floor = 'N/A'
            room = 'N/A'
            if '_f' in device and 'r' in device:
                try:
                    parts = device.split('_')
                    for p in parts:
                        if p.startswith('f') and 'r' in p:
                            floor = p.split('r')[0].replace('f', '')
                            room = p.split('r')[1]
                            break
                except:
                    pass
            
            # Calculate per-device packet loss
            expected = data['max_seq'] + 1  # seq starts from 0
            received = len(data['received_seq'])
            lost = expected - received
            loss_rate = (lost / expected * 100) if expected > 0 else 0
            
            summary['device_stats'].append({
                'device': device,
                'floor': floor,
                'room': room,
                'type': data['type'],
                'count': len(data['delays']),
                'avg_delay': mean(data['delays']),
                'min_delay': min(data['delays']),
                'max_delay': max(data['delays']),
                'expected': expected,
                'received': received,
                'lost': lost,
                'loss_rate': loss_rate
            })
    
    # Sort by avg delay descending
    summary['device_stats'].sort(key=lambda x: x['avg_delay'], reverse=True)

    return summary

def print_summary(summary):
    """Print summary to console"""
    print("\n" + "=" * 70)
    print(" " * 20 + "SIMULATION SUMMARY")
    print("=" * 70 + "\n")

    for msg_type in ['anomaly', 'normal']:
        if msg_type in summary and summary[msg_type]:
            s = summary[msg_type]
            loss = s['loss_info']
            print(f"{msg_type.upper()}:")
            print(f"  Devices           : {s['num_devices']}")
            print(f"  Messages Received : {s['count']}")
            print(f"  Avg Delay         : {s['avg_delay']:.2f} ms")
            print(f"  Min Delay         : {s['min_delay']:.2f} ms")
            print(f"  Max Delay         : {s['max_delay']:.2f} ms")
            print(f"  Std Dev Delay     : {s['std_delay']:.2f} ms")
            print(f"  Avg Jitter        : {s['avg_jitter']:.2f} ms")
            print(f"\n  PACKET LOSS:")
            print(f"    Expected        : {loss['expected']} messages")
            print(f"    Received        : {loss['received']} messages")
            print(f"    Lost            : {loss['lost']} messages")
            print(f"    Loss Rate       : {loss['loss_rate']:.2f}%")
            print(f"    Avg Loss/Device : {loss['avg_loss_per_device']:.2f}%")
            print()

    print(f"TOTAL:")
    print(f"  Duration          : {summary['total']['duration']:.2f} s")
    print(f"  Total Messages    : {summary['total']['total_messages']}")
    print(f"  Throughput        : {summary['total']['throughput']:.2f} msg/s")
    print()

    # Per-sensor metrics
    if 'device_stats' in summary and summary['device_stats']:
        print("=" * 90)
        print("PER-SENSOR METRICS (sorted by avg delay, highest first)")
        print("=" * 90)
        print(f"\n{'Sensor':<26} {'Floor':>5} {'Type':<8} {'Recv':>6} {'Expect':>6} {'Loss':>7} {'Avg Delay':>12}")
        print(f"{'-'*26} {'-'*5} {'-'*8} {'-'*6} {'-'*6} {'-'*7} {'-'*12}")
        
        for ds in summary['device_stats']:
            loss_str = f"{ds['loss_rate']:.1f}%"
            print(f"{ds['device']:<26} {ds['floor']:>5} {ds['type']:<8} {ds['received']:>6} {ds['expected']:>6} {loss_str:>7} {ds['avg_delay']:>10.2f}ms")
        
        # Per-floor summary
        floor_data = defaultdict(lambda: {
            'anomaly_delays': [], 'normal_delays': [], 
            'anomaly_loss': [], 'normal_loss': [],
            'total_count': 0, 'total_lost': 0
        })
        for ds in summary['device_stats']:
            floor_data[ds['floor']][f"{ds['type']}_delays"].append(ds['avg_delay'])
            floor_data[ds['floor']][f"{ds['type']}_loss"].append(ds['loss_rate'])
            floor_data[ds['floor']]['total_count'] += ds['received']
            floor_data[ds['floor']]['total_lost'] += ds['lost']
        
        print(f"\nPER-FLOOR SUMMARY:")
        print(f"{'-'*70}")
        print(f"{'Floor':>5} {'Messages':>10} {'Lost':>6} {'Anomaly Avg':>14} {'Normal Avg':>14} {'Loss%':>8}")
        print(f"{'-'*5} {'-'*10} {'-'*6} {'-'*14} {'-'*14} {'-'*8}")
        
        for floor in sorted(floor_data.keys()):
            fd = floor_data[floor]
            anomaly_avg = mean(fd['anomaly_delays']) if fd['anomaly_delays'] else 0
            normal_avg = mean(fd['normal_delays']) if fd['normal_delays'] else 0
            total = fd['total_count'] + fd['total_lost']
            loss_pct = (fd['total_lost'] / total * 100) if total > 0 else 0
            print(f"{floor:>5} {fd['total_count']:>10} {fd['total_lost']:>6} {anomaly_avg:>12.2f}ms {normal_avg:>12.2f}ms {loss_pct:>7.2f}%")
        print()

def save_summary(summary, output_file):
    """Save summary to file"""
    with open(output_file, 'w') as f:
        f.write("=" * 70 + "\n")
        f.write(" " * 20 + "SIMULATION SUMMARY\n")
        f.write("=" * 70 + "\n\n")

        for msg_type in ['anomaly', 'normal']:
            if msg_type in summary and summary[msg_type]:
                s = summary[msg_type]
                loss = s['loss_info']
                f.write(f"{msg_type.upper()}:\n")
                f.write(f"  Devices           : {s['num_devices']}\n")
                f.write(f"  Messages Received : {s['count']}\n")
                f.write(f"  Avg Delay         : {s['avg_delay']:.2f} ms\n")
                f.write(f"  Min Delay         : {s['min_delay']:.2f} ms\n")
                f.write(f"  Max Delay         : {s['max_delay']:.2f} ms\n")
                f.write(f"  Std Dev Delay     : {s['std_delay']:.2f} ms\n")
                f.write(f"  Avg Jitter        : {s['avg_jitter']:.2f} ms\n")
                f.write(f"\n  PACKET LOSS:\n")
                f.write(f"    Expected        : {loss['expected']} messages\n")
                f.write(f"    Received        : {loss['received']} messages\n")
                f.write(f"    Lost            : {loss['lost']} messages\n")
                f.write(f"    Loss Rate       : {loss['loss_rate']:.2f}%\n")
                f.write(f"    Avg Loss/Device : {loss['avg_loss_per_device']:.2f}%\n")
                f.write("\n")

        f.write(f"TOTAL:\n")
        f.write(f"  Duration          : {summary['total']['duration']:.2f} s\n")
        f.write(f"  Total Messages    : {summary['total']['total_messages']}\n")
        f.write(f"  Throughput        : {summary['total']['throughput']:.2f} msg/s\n")

        # Per-sensor metrics
        if 'device_stats' in summary and summary['device_stats']:
            f.write("\n" + "=" * 90 + "\n")
            f.write("PER-SENSOR METRICS (sorted by avg delay, highest first)\n")
            f.write("=" * 90 + "\n\n")
            f.write(f"{'Sensor':<26} {'Floor':>5} {'Type':<8} {'Recv':>6} {'Expect':>6} {'Loss':>7} {'Avg Delay':>12}\n")
            f.write(f"{'-'*26} {'-'*5} {'-'*8} {'-'*6} {'-'*6} {'-'*7} {'-'*12}\n")
            
            for ds in summary['device_stats']:
                loss_str = f"{ds['loss_rate']:.1f}%"
                f.write(f"{ds['device']:<26} {ds['floor']:>5} {ds['type']:<8} {ds['received']:>6} {ds['expected']:>6} {loss_str:>7} {ds['avg_delay']:>10.2f}ms\n")
            
            # Per-floor summary
            floor_data = defaultdict(lambda: {
                'anomaly_delays': [], 'normal_delays': [], 
                'anomaly_loss': [], 'normal_loss': [],
                'total_count': 0, 'total_lost': 0
            })
            for ds in summary['device_stats']:
                floor_data[ds['floor']][f"{ds['type']}_delays"].append(ds['avg_delay'])
                floor_data[ds['floor']][f"{ds['type']}_loss"].append(ds['loss_rate'])
                floor_data[ds['floor']]['total_count'] += ds['received']
                floor_data[ds['floor']]['total_lost'] += ds['lost']
            
            f.write(f"\nPER-FLOOR SUMMARY:\n")
            f.write(f"{'-'*70}\n")
            f.write(f"{'Floor':>5} {'Messages':>10} {'Lost':>6} {'Anomaly Avg':>14} {'Normal Avg':>14} {'Loss%':>8}\n")
            f.write(f"{'-'*5} {'-'*10} {'-'*6} {'-'*14} {'-'*14} {'-'*8}\n")
            
            for floor in sorted(floor_data.keys()):
                fd = floor_data[floor]
                anomaly_avg = mean(fd['anomaly_delays']) if fd['anomaly_delays'] else 0
                normal_avg = mean(fd['normal_delays']) if fd['normal_delays'] else 0
                total = fd['total_count'] + fd['total_lost']
                loss_pct = (fd['total_lost'] / total * 100) if total > 0 else 0
                f.write(f"{floor:>5} {fd['total_count']:>10} {fd['total_lost']:>6} {anomaly_avg:>12.2f}ms {normal_avg:>12.2f}ms {loss_pct:>7.2f}%\n")

    print(f"Summary saved to: {output_file}")

def save_summary_excel(summary, output_file):
    """Save summary to Excel file with formatting"""
    if not EXCEL_AVAILABLE:
        print("Skipping Excel export (openpyxl not installed)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics Summary"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    metric_font = Font(size=10)
    value_font = Font(size=10, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "MQTT-SDN METRICS SUMMARY"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    row += 2

    # Anomaly Section
    if 'anomaly' in summary and summary['anomaly']:
        s = summary['anomaly']
        loss = s['loss_info']

        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "ANOMALY TRAFFIC"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Delay Metrics
        metrics_data = [
            ("Devices", s['num_devices'], ""),
            ("Messages Received", s['count'], "messages"),
            ("Avg Delay", f"{s['avg_delay']:.2f}", "ms"),
            ("Min Delay", f"{s['min_delay']:.2f}", "ms"),
            ("Max Delay", f"{s['max_delay']:.2f}", "ms"),
            ("Std Dev Delay", f"{s['std_delay']:.2f}", "ms"),
            ("Avg Jitter", f"{s['avg_jitter']:.2f}", "ms"),
        ]

        for metric, value, unit in metrics_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'A{row}'].font = metric_font
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "PACKET LOSS"
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        loss_data = [
            ("Expected", loss['expected'], "messages"),
            ("Received", loss['received'], "messages"),
            ("Lost", loss['lost'], "messages"),
            ("Loss Rate", f"{loss['loss_rate']:.2f}", "%"),
            ("Avg Loss/Device", f"{loss['avg_loss_per_device']:.2f}", "%"),
        ]

        for metric, value, unit in loss_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'A{row}'].font = metric_font
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        row += 2

    # Normal Section
    if 'normal' in summary and summary['normal']:
        s = summary['normal']
        loss = s['loss_info']

        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "NORMAL TRAFFIC"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Delay Metrics
        metrics_data = [
            ("Devices", s['num_devices'], ""),
            ("Messages Received", s['count'], "messages"),
            ("Avg Delay", f"{s['avg_delay']:.2f}", "ms"),
            ("Min Delay", f"{s['min_delay']:.2f}", "ms"),
            ("Max Delay", f"{s['max_delay']:.2f}", "ms"),
            ("Std Dev Delay", f"{s['std_delay']:.2f}", "ms"),
            ("Avg Jitter", f"{s['avg_jitter']:.2f}", "ms"),
        ]

        for metric, value, unit in metrics_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'A{row}'].font = metric_font
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "PACKET LOSS"
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        loss_data = [
            ("Expected", loss['expected'], "messages"),
            ("Received", loss['received'], "messages"),
            ("Lost", loss['lost'], "messages"),
            ("Loss Rate", f"{loss['loss_rate']:.2f}", "%"),
            ("Avg Loss/Device", f"{loss['avg_loss_per_device']:.2f}", "%"),
        ]

        for metric, value, unit in loss_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'A{row}'].font = metric_font
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        row += 2

    # Total Section
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "TOTAL"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    total_data = [
        ("Duration", f"{summary['total']['duration']:.2f}", "s"),
        ("Total Messages", summary['total']['total_messages'], "messages"),
        ("Throughput", f"{summary['total']['throughput']:.2f}", "msg/s"),
    ]

    for metric, value, unit in total_data:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = metric_font
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12

    # Save workbook
    wb.save(output_file)
    print(f"Excel summary saved to: {output_file}")

def save_summary_csv(summary, output_file):
    """Save summary to CSV file (Excel-compatible)"""
    import csv as csv_module
    with open(output_file, 'w', newline='') as f:
        writer = csv_module.writer(f)

        # Title
        writer.writerow(['MQTT-SDN METRICS SUMMARY'])
        writer.writerow([])

        # Anomaly Section
        if 'anomaly' in summary and summary['anomaly']:
            s = summary['anomaly']
            loss = s['loss_info']
            writer.writerow(['ANOMALY TRAFFIC'])
            writer.writerow(['Metric', 'Value', 'Unit'])
            writer.writerow(['Devices', s['num_devices'], ''])
            writer.writerow(['Messages Received', s['count'], 'messages'])
            writer.writerow(['Avg Delay', f"{s['avg_delay']:.2f}", 'ms'])
            writer.writerow(['Min Delay', f"{s['min_delay']:.2f}", 'ms'])
            writer.writerow(['Max Delay', f"{s['max_delay']:.2f}", 'ms'])
            writer.writerow(['Std Dev Delay', f"{s['std_delay']:.2f}", 'ms'])
            writer.writerow(['Avg Jitter', f"{s['avg_jitter']:.2f}", 'ms'])
            writer.writerow([])
            writer.writerow(['PACKET LOSS'])
            writer.writerow(['Expected', loss['expected'], 'messages'])
            writer.writerow(['Received', loss['received'], 'messages'])
            writer.writerow(['Lost', loss['lost'], 'messages'])
            writer.writerow(['Loss Rate', f"{loss['loss_rate']:.2f}", '%'])
            writer.writerow(['Avg Loss/Device', f"{loss['avg_loss_per_device']:.2f}", '%'])
            writer.writerow([])

        # Normal Section
        if 'normal' in summary and summary['normal']:
            s = summary['normal']
            loss = s['loss_info']
            writer.writerow(['NORMAL TRAFFIC'])
            writer.writerow(['Metric', 'Value', 'Unit'])
            writer.writerow(['Devices', s['num_devices'], ''])
            writer.writerow(['Messages Received', s['count'], 'messages'])
            writer.writerow(['Avg Delay', f"{s['avg_delay']:.2f}", 'ms'])
            writer.writerow(['Min Delay', f"{s['min_delay']:.2f}", 'ms'])
            writer.writerow(['Max Delay', f"{s['max_delay']:.2f}", 'ms'])
            writer.writerow(['Std Dev Delay', f"{s['std_delay']:.2f}", 'ms'])
            writer.writerow(['Avg Jitter', f"{s['avg_jitter']:.2f}", 'ms'])
            writer.writerow([])
            writer.writerow(['PACKET LOSS'])
            writer.writerow(['Expected', loss['expected'], 'messages'])
            writer.writerow(['Received', loss['received'], 'messages'])
            writer.writerow(['Lost', loss['lost'], 'messages'])
            writer.writerow(['Loss Rate', f"{loss['loss_rate']:.2f}", '%'])
            writer.writerow(['Avg Loss/Device', f"{loss['avg_loss_per_device']:.2f}", '%'])
            writer.writerow([])

        # Total Section
        writer.writerow(['TOTAL'])
        writer.writerow(['Metric', 'Value', 'Unit'])
        writer.writerow(['Duration', f"{summary['total']['duration']:.2f}", 's'])
        writer.writerow(['Total Messages', summary['total']['total_messages'], 'messages'])
        writer.writerow(['Throughput', f"{summary['total']['throughput']:.2f}", 'msg/s'])

    print(f"CSV summary saved to: {output_file} (open with Excel)")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 generate_summary_manual_v2.py <path_to_mqtt_metrics_log.csv>")
        print("\nExample:")
        print("  python3 generate_summary_manual_v2.py /home/mqtt-sdn/results/06-dscp-qos-13switches/run_2025-11-24_13-36-51/mqtt_metrics_log.csv")
        sys.exit(1)

    csv_file = sys.argv[1]

    if not os.path.exists(csv_file):
        print(f"Error: File not found: {csv_file}")
        sys.exit(1)

    # Calculate summary
    summary = calculate_summary(csv_file)

    # Print to console
    print_summary(summary)

    # Save to file in same directory as CSV
    output_dir = os.path.dirname(csv_file)

    # Save text summary
    output_txt = os.path.join(output_dir, "metrics_summary_v2.txt")
    save_summary(summary, output_txt)

    # Save Excel summary (if openpyxl available)
    if EXCEL_AVAILABLE:
        output_xlsx = os.path.join(output_dir, "metrics_summary_v2.xlsx")
        save_summary_excel(summary, output_xlsx)

    # Save CSV summary (always available, Excel-compatible)
    output_csv = os.path.join(output_dir, "metrics_summary_v2_excel.csv")
    save_summary_csv(summary, output_csv)

    print("\n" + "=" * 70)
    print("Done!")
    print("=" * 70)
